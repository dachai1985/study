# -*- coding: utf-8 -*-

import os
import sys
import pandas as pd
import win32com.client as win32
import shutil
from COMMON import config
import openpyxl

names_to_count = ['张三', '李四']  # 需要统计的名字列表  
tc_count_flag = False
error_file_count = 0

# 获取当前folder内所有excel文件名
def get_excel_name_from_floder(folder_path):
    file_name = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if os.path.splitext(file)[1] == '.xlsx':
                file_name.append(os.path.join(root, file))
    return file_name

# 撤销当前Excel保护
def unlock_excel(file_name, password):  
    # 创建Excel应用程序实例  
    # excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = False  # 设置为不可见
    excel.DisplayAlerts = False  # 关闭警告提示
    try:  
        # 打开工作簿  
        workbook = excel.Workbooks.Open(file_name, Password=password)    
        # 如果有工作表保护，解除保护  
        for sheet in workbook.Sheets:  
            if sheet.ProtectContents:
                sheet.Unprotect(Password=password)  
                # 保存工作簿
                workbook.Save()
        workbook.Close()
    except Exception as e:  
        print(f"Error unlocking or reading the file: {e}", '\n')

# 移动文件到pass folder文件夹
def move_to_pass_folder(file_name):
    pass_folder = os.path.join(os.path.dirname(file_name), "pass folder")
    if not os.path.exists(pass_folder):
        # 如果不存在，则创建文件夹  
        os.makedirs(pass_folder)
    # 获取源文件的基本名称和扩展名  
    base_name, extension = os.path.splitext(os.path.basename(file_name))
    # 构建目标文件的初始路径  
    target_file = os.path.join(pass_folder, base_name + extension)
    # 检查目标文件是否存在,如果存在，生成一个新的文件名   
    if os.path.exists(file_name):
        counter = 1 
        while os.path.exists(target_file):  
            # 在基本名称后添加序号以生成新文件名  
            new_base_name = f"{base_name}_{counter}"  
            target_file = os.path.join(pass_folder, new_base_name + extension)  
            counter += 1
    shutil.move(file_name, target_file)

# 获取当前Excel所有sheet列表
def get_excel_sheet_list_pyxl(file_name):
    try:  
        # 加载Excel工作簿  
        workbook = openpyxl.load_workbook(filename=file_name)  
        # 获取所有工作表名称  
        sheet_names = workbook.sheetnames  
        return sheet_names  
    except Exception as e:  
        print(f"Failed to read {file_name}")  
        print(f"Error: {e}")  
        return []
    
# 定义函数来执列查找操作，涵盖合并单元格的内容 
def find_position_with_string(sheet, search_string):
    # 获取所有合并单元格的范围  
    merged_cells_ranges = sheet.merged_cells.ranges  
      
    # 建立一个字典来映射合并单元格的起始单元格到其内容  
    merged_cell_values = {}  
    for merged_range in merged_cells_ranges:  
        min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(str(merged_range))  
        top_left_cell = sheet.cell(row=min_row, column=min_col)  
        merged_cell_values[(min_row, min_col)] = top_left_cell.value
      
    # 遍历所有列，查找包含搜索字符串的列  
    for column_index, column in enumerate(sheet.columns, start=1):  
        for cell in column:
            # 获取单元格的坐标  
            row, col = cell.row, cell.column  
              
            # 检查单元格是否在合并单元格范围内  
            if (row, col) in merged_cell_values:  
                cell_value = merged_cell_values[(row, col)]
            else:  
                cell_value = cell.value

            # print(repr(cell.value))  # 这将显示包含换行符的原始字符串表示

            # 检查单元格的值是否包含搜索字符串  
            if cell_value and search_string.lower() in str(cell_value).lower():
                print(repr(cell.value))  # 这将显示包含换行符的原始字符串表示
                # 如果找到，返回列索引  
                return column_index, row
      
    # 如果没有找到，返回None  
    return None 

# 统计当前sheet 各tester count
def get_tester_count_sheet(file_name, sheet_name, names_to_count):
    workbook = openpyxl.load_workbook(file_name)  
    sheet = workbook[sheet_name]   

    column_name_to_check = 'Tester'
    position = find_position_with_string(sheet, column_name_to_check)

    global tc_count_flag
    if position is not None:  
        print(f"Tester found in the sheet {sheet_name} of column: {position}")  
        # 初始化一个字典来存储每个名字的计数  
        count_dict = {name: 0 for name in names_to_count} 
        column_index = 1  # Excel中列索引从1开始，A列是1  
        # 遍历指定列中的每一行  
        for row in sheet.iter_rows(min_col=column_index, max_col=column_index, max_row=sheet.max_row):  
            for cell in row:  
                # 去除单元格值的前后空格，并检查是否在我们要统计的名字列表中  
                cell_value_stripped = str(cell.value).strip() if cell.value is not None else ''  
                if cell_value_stripped in names_to_count:  
                    count_dict[cell_value_stripped] += 1  

        for name, count in count_dict.items():  
            print(f"{name} 出现了 {count} 次")  
        tc_count_flag = True
    else:  
        print(f"Tester not found in the sheet {sheet_name}")  
        tester_count_sheet = 0
    return tester_count_sheet

# 统计当前sheet creator count
def get_creator_count_sheet(file_name, sheet_name, names_to_count):
    workbook = openpyxl.load_workbook(file_name)  
    sheet = workbook[sheet_name]   

    column_name_to_check = 'Creator'
    position = find_position_with_string(sheet, column_name_to_check)

    global tc_count_flag
    if position is not None:  
        print(f"Creator found in the sheet {sheet_name} of column: {position}")  
        # 初始化一个字典来存储每个名字的计数  
        count_dict = {name: 0 for name in names_to_count} 
        column_index = 1  # Excel中列索引从1开始，A列是1  
        # 遍历指定列中的每一行  
        for row in sheet.iter_rows(min_col=column_index, max_col=column_index, max_row=sheet.max_row):  
            for cell in row:  
                # 去除单元格值的前后空格，并检查是否在我们要统计的名字列表中  
                cell_value_stripped = str(cell.value).strip() if cell.value is not None else ''  
                if cell_value_stripped in names_to_count:  
                    count_dict[cell_value_stripped] += 1    

        for name, count in count_dict.items():  
            print(f"{name} 出现了 {count} 次")  
        tc_count_flag = True
    else:  
        print(f"Creator not found in the sheet {sheet_name}")  
    return count_dict

 # 获取当前Excel file BTS Link列非空数值
def get_count_file(file_name):
    bts_count_file = 0
    global error_file_count
    global tc_count_flag
    try:
        print('filename==========', file_name)
        unlock_excel(file_name,'')
        sheet_names = get_excel_sheet_list_pyxl(file_name)
        print("sheet_names=======", sheet_names)
        tc_count_flag = False

        # 初始化一个字典来存储所有sheet的汇总结果  
        total_counts_file = {name: 0 for name in names_to_count}

        for sheet_name in sheet_names:
            sheet_counts  = get_tester_count_sheet(file_name, sheet_name, names_to_count)
            for name, count in sheet_counts.items():  
                total_counts_file[name] += count
        print(f"{file_name} file bts count========== {bts_count_file}", '\n')
        if tc_count_flag:
            move_to_pass_folder(file_name)    
        else:
            error_file_count += 1
    except Exception as e:
        #print(f"Failed to read {file_name}. Moving to {error_folder}...") 
        #shutil.move(file_name, error_folder)    
        print(f"Error: {e}", '\n')
        error_file_count += 1
    return total_counts_file

# 遍历目录下的所有game文件夹,统计各文件的BTS count
def get_count_game(reading_folder):
    # 字典用于存储每个文件夹的数据  
    bts_count_game = {}
    # 要跳过的子文件夹名称  
    skip_folder_name = "pass folder" 

    for folder_name, subfolders, filenames in os.walk(reading_folder):
        # 检查当前文件夹是否是我们要跳过的文件夹  
        if skip_folder_name in folder_name:
            continue  # 如果是，跳过当前循环，即不处理该文件夹

        # 初始化一个字典来存储所有文件的汇总结果  
        total_counts = {name: 0 for name in names_to_count}  
        
        # 过滤出Excel文件（假设文件扩展名为.xlsx
        excel_files = [f for f in filenames if f.endswith('.xlsx')]
        # 如果文件夹中有Excel文件  
        if excel_files:
            # 初始化该文件夹的数据列表
            bts_count_game[folder_name] = []
            # 遍历该文件夹下的所有Excel文件  
            for excel_file in excel_files:  
                file_path = os.path.join(folder_name, excel_file)
                try:
                    # 读取Excel文件  
                    file_counts = get_count_file(file_path)
                    
                     # 将当前文件的统计结果合并到总结果中  
                    for name, count in file_counts.items():  
                        total_counts[name] += count
                except Exception as e:  
                    print(f"Error reading {file_path}: {e}", '\n')
    return total_counts

if __name__=='__main__':
    reading_folder = r'C:\\Users\\win\\Desktop\\test\\tantan'
    bts_count_game_tc_key = get_count_game(reading_folder)
    print ('bts_count_game_tc_key==========', bts_count_game_tc_key)
    print ('error_file_count========', error_file_count)