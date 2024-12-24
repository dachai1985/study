import openpyxl
import os
from Pytest_API.utils.LogUtil import my_log
from typing import List, Dict

# 参数化 pytest list
class ExcelReader:
    def __init__(self, file_path, sheet_name=None):
        self.log = my_log()
        self.log.info(f"Checking for file at path: {file_path}")  # 记录日志
        # 验证excel文件是否存在
        if os.path.exists(file_path):
            self.file_path = file_path
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.sheet_name = sheet_name if sheet_name else self.workbook.sheetnames[0]
            self._data_list = []
        else:
            raise FileNotFoundError(f"File not found at path: {file_path}")

    # 读取excel文件 返回list 元素:字典
    # [{"a":"a1","b":"b1"},{"a":"a2","b":"b2"}]
    # 获取首行的信息
    def read_test_data(self) -> List[Dict[str, str]]:

        # 数据列表存在不读取，不存在读取
        if not self._data_list:
            try:
                # 获取第一个sheet
                sheet = self.workbook[self.sheet_name]
                # 获取首行信息
                head = [cell.value for cell in sheet[1]]
                # 遍历测试行，与首行组成dict，放在list
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # 列1的值为None，跳过
                    if row[0] is None:
                        continue
                    data_dict = dict(zip(head, row))
                    self._data_list.append(data_dict)
            except KeyError as e:
                self.log.error(f"Sheet not found: {e}")
                raise
            except Exception as e:
                self.log.error(f"An error occurred while reading the Excel file: {e}")
                raise

        return self._data_list


# 结果返回
# def data_dict_test():
#     head = ["a", "b"]
#     value1 = ["a1", "b1"]
#     value2 = ["a2", "b2"]
#     data_list = list()
#     data_list.append(dict(zip(head, value1)))
#     data_list.append(dict(zip(head, value2)))
#     print(data_list)

if __name__ == '__main__':
    # data_dict_test()
    excel_reader = ExcelReader("../data/api tc.xlsx", "Sheet1")
    data_list = excel_reader.read_test_data()
    print(data_list)

